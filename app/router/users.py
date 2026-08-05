import csv
import io
import secrets
from typing import List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.dependencies import CurrentUserDep, CurrentSuperAdminDep, CurrentBulkImportDep, DBSessionDep
from app.core.tenant import add_org_filter, get_current_org_id, is_platform
from app.models.user import User
from app.models.user_role import UserRole
from app.models.role import Role
from app.models.department import Department
from app.models.enums import RoleName, UserStatus
from app.schema.user import UserCreate, UserRead, UserUpdate, UserBulkImport, UserInviteResponse
from app.core.security import get_password_hash
from pydantic import BaseModel

router = APIRouter(prefix="/api/users", tags=["users"])


@router.post("/invite", response_model=UserInviteResponse, status_code=status.HTTP_201_CREATED)
async def invite_user(payload: UserCreate, current_user: CurrentSuperAdminDep, db: DBSessionDep):
    result = await db.execute(select(User).where(User.email == payload.email))
    if result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User with this email already exists")
    if payload.organization_id != current_user.organization_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot invite users to another organization")
    generated_password = secrets.token_urlsafe(12)
    user = User(
        **payload.model_dump(exclude={'status'}),
        hashed_password=get_password_hash(generated_password),
        status=UserStatus.active,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return UserInviteResponse(
        id=user.id,
        email=user.email,
        first_name=user.first_name,
        last_name=user.last_name,
        organization_id=user.organization_id,
        department_id=user.department_id,
        status=user.status,
        mfa_enabled=user.mfa_enabled,
        sso_provider=user.sso_provider,
        created_at=user.created_at,
        last_login_at=user.last_login_at,
        roles=[],
        generated_password=generated_password,
    )


@router.get("/", response_model=List[UserRead])
async def list_users(current_user: CurrentUserDep, db: DBSessionDep) -> List[User]:
    result = await db.execute(
        select(User).where(User.id == current_user.id).options(selectinload(User.roles))
    )
    user = result.scalar_one_or_none()
    is_admin = user and any(
        role.name in (RoleName.super_admin, RoleName.compliance_admin) for role in user.roles
    )
    if is_admin:
        query = select(User).where(User.organization_id == current_user.organization_id).options(selectinload(User.roles))
        result = await db.execute(query)
        return result.scalars().all()
    else:
        return [user]


@router.get("/{user_id}", response_model=UserRead)
async def get_user_profile(user_id: int, current_user: CurrentUserDep, db: DBSessionDep) -> User:
    result = await db.execute(
        select(User).where(User.id == user_id, User.organization_id == current_user.organization_id).options(selectinload(User.roles))
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return user


@router.patch("/{user_id}", response_model=UserRead)
async def update_user_profile(user_id: int, payload: UserUpdate, current_user: CurrentSuperAdminDep, db: DBSessionDep) -> User:
    result = await db.execute(
        select(User).where(User.id == user_id, User.organization_id == current_user.organization_id).options(selectinload(User.roles))
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(user, key, value)
    u_id = user.id
    await db.commit()
    result = await db.execute(select(User).where(User.id == u_id).options(selectinload(User.roles)))
    return result.scalar_one()


@router.post("/{user_id}/deactivate", status_code=status.HTTP_200_OK)
async def deactivate_user(user_id: int, current_user: CurrentSuperAdminDep, db: DBSessionDep):
    result = await db.execute(
        select(User).where(User.id == user_id, User.organization_id == current_user.organization_id)
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    user.status = UserStatus.deactivated
    await db.commit()
    return {"message": "User deactivated successfully"}


@router.post("/bulk-import", status_code=status.HTTP_201_CREATED)
async def bulk_import_users(
    current_user: CurrentBulkImportDep,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    org_id = current_user.organization_id

    # Read roles & departments for mapping
    roles_result = await db.execute(
        select(Role).where((Role.organization_id == org_id) | (Role.organization_id.is_(None)))
    )
    roles_map = {r.name.lower(): r.id for r in roles_result.scalars().all()}

    depts_result = await db.execute(
        select(Department).where(Department.organization_id == org_id)
    )
    depts_map = {d.name.lower(): d.id for d in depts_result.scalars().all()}

    # Read file content
    content = await file.read()
    rows = []

    if file.filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{k.lower(): v for k, v in row.items()} for row in reader]
    elif file.filename.endswith((".xls", ".xlsx")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        header = [cell.value.strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=ws.max_row))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(header, [str(v) if v is not None else "" for v in row])))
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file format. Use CSV or Excel (.xlsx)")

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty or has no data rows")

    # Validate headers
    expected_cols = {"name", "first_name", "last_name", "email", "role", "department"}
    col_set = {k.lower() for k in rows[0].keys()}
    name_col = next((c for c in ("name", "first_name", "last_name") if c in col_set), None)
    if not name_col or "email" not in col_set or "role" not in col_set:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV/Excel must include columns: Name (or First Name + Last Name), Email, Role, Department (optional)"
        )

    created = []
    errors = []

    for i, row in enumerate(rows, start=2):
        email = row.get("email", "").strip().lower()
        role_name = row.get("role", "").strip()
        dept_name = row.get("department", "").strip()

        # Parse name
        if "name" in col_set:
            parts = row.get("name", "").strip().split(" ", 1)
            first_name = parts[0].strip() if parts else ""
            last_name = parts[1].strip() if len(parts) > 1 else ""
        else:
            first_name = row.get("first_name", "").strip()
            last_name = row.get("last_name", "").strip()

        line_errs = []

        if not email:
            line_errs.append("Missing email")
        if not first_name:
            line_errs.append("Missing first name")
        if not last_name:
            line_errs.append("Missing last name")
        if not role_name:
            line_errs.append("Missing role")

        if line_errs:
            errors.append({"row": i, "email": email or f"row {i}", "errors": line_errs})
            continue

        # Check duplicate email
        existing = await db.execute(select(User).where(User.email == email))
        if existing.scalar_one_or_none():
            errors.append({"row": i, "email": email, "errors": ["Email already exists"]})
            continue

        # Map role
        role_id = roles_map.get(role_name.lower())
        if not role_id:
            errors.append({"row": i, "email": email, "errors": [f"Role '{role_name}' not found"]})
            continue

        # Map department (optional)
        dept_id = depts_map.get(dept_name.lower()) if dept_name else None

        # Create user
        password = secrets.token_urlsafe(12)
        user = User(
            email=email,
            first_name=first_name,
            last_name=last_name,
            organization_id=org_id,
            department_id=dept_id,
            status=UserStatus.invited,
            hashed_password=get_password_hash(password),
        )
        db.add(user)
        await db.flush()

        # Assign role
        db.add(UserRole(user_id=user.id, role_id=role_id))

        created.append({
            "id": user.id,
            "email": email,
            "first_name": first_name,
            "last_name": last_name,
            "generated_password": password,
        })

    await db.commit()

    return {
        "success_count": len(created),
        "error_count": len(errors),
        "created": created,
        "errors": errors,
    }


@router.delete("/{user_id}", status_code=status.HTTP_200_OK)
async def delete_user(user_id: int, current_user: CurrentSuperAdminDep, db: DBSessionDep):
    result = await db.execute(
        select(User).where(User.id == user_id, User.organization_id == current_user.organization_id)
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    await db.delete(user)
    await db.commit()
    return {"message": "User deleted successfully"}


# Role Assignment
class RoleAssignment(BaseModel):
    user_id: int
    role_id: int


class RoleSync(BaseModel):
    user_id: int
    role_ids: List[int]


async def _assert_user_in_org(user_id: int, org_id: int, db: AsyncSession) -> User:
    result = await db.execute(select(User).where(User.id == user_id, User.organization_id == org_id))
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found in your organization")
    return user


@router.post("/sync-roles", status_code=status.HTTP_200_OK)
async def sync_user_roles(payload: RoleSync, current_user: CurrentSuperAdminDep, db: DBSessionDep):
    await _assert_user_in_org(payload.user_id, current_user.organization_id, db)
    await db.execute(delete(UserRole).where(UserRole.user_id == payload.user_id))
    new_roles = [UserRole(user_id=payload.user_id, role_id=r_id) for r_id in payload.role_ids]
    db.add_all(new_roles)
    await db.commit()
    return {"message": "Roles synced successfully"}


@router.post("/assign-role", status_code=status.HTTP_201_CREATED)
async def assign_role(payload: RoleAssignment, current_user: CurrentSuperAdminDep, db: DBSessionDep):
    await _assert_user_in_org(payload.user_id, current_user.organization_id, db)
    user_role = UserRole(user_id=payload.user_id, role_id=payload.role_id)
    db.add(user_role)
    await db.commit()
    return {"message": "Role assigned successfully"}
